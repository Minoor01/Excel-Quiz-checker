import openpyxl 
quizfile = openpyxl.load_workbook("directory") 
sheet1 = quizfile["Form Responses 1"]
columns = sheet1.max_column+1
rows = sheet1.max_row+1
count = 0
print(sheet1.cell(2,3).value)
for row in range(2,rows):
    for column in range (7,columns+1):
        if column == 7 and sheet1.cell(row,column).value == "Correct1":
            count += 1
        if column == 8 and sheet1.cell(row,column).value == "Correct1":
            count += 1
        if column == 9 and sheet1.cell(row,column).value == "Correct1":
            count += 1
        if column == 10 and sheet1.cell(row,column).value == "Correct1":
            count += 1
        if column == 11 and sheet1.cell(row,column).value == "Correct1":
            count += 1
        if column == 12 and sheet1.cell(row,column).value == "Correct1":
            count += 1
        if column == 13 and sheet1.cell(row,column).value == "Correct1":
            count += 1
        if column == 14 and sheet1.cell(row,column).value == "Correct1":
            count += 1
        if column == 15 and sheet1.cell(row,column).value == "Correct1":
            count += 1
        if column == 16 and sheet1.cell(row,column).value == "Correct1":
            count += 1
        if column == 17:
            sheet1.cell(row,column).value = "{}/10".format(count)
    count = 0


quizfile.save("Directory")